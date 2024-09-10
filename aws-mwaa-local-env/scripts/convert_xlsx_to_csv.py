import csv
import logging
import os

from openpyxl import load_workbook

def convert_xlsx_to_csv(xlsx_path, csv_dir):
    if not os.path.exists(csv_dir):
        os.makedirs(csv_dir)

    wb = load_workbook(xlsx_path, data_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        csv_filepath = f"{csv_dir}/{sheet}.csv"
        with open(csv_filepath, 'w', newline='') as f:
            c = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                #  Exclude empty rows
                if any(value is not None for value in row):
                    c.writerow(row)
    logging.info(f"""
        Successfully converted XLSX file to CSVs.
        Source XLSX PATH: '{xlsx_path}'
        CSV destination directory: '{csv_dir}'
    """)
